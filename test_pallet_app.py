import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import datetime, tempfile, os, math

# ── PALETTE ───────────────────────────────────────────────────────────────────
NAVY        = "0D2137"
NAVY_MID    = "1A3A5C"
NAVY_LIGHT  = "1E4D78"
STEEL_BLUE  = "2E6DA4"
ORANGE      = "E87722"
ORANGE_DARK = "B85C0A"
ORANGE_LIGHT= "FEF0E4"
BLUE_LIGHT  = "EAF2FB"
CREAM       = "FAFAFA"
WHITE       = "FFFFFF"
ROW_ALT     = "F0F6FC"
TABLE_DARK  = "D6E8F5"
SILVER_LINE = "B8CFE0"
DARK_TEXT   = "0D2137"
MID_TEXT    = "2E5070"
LIGHT_TEXT  = "6A8FAA"

# ── BORDER HELPERS ────────────────────────────────────────────────────────────
def side(color, style="thin"):
    return Side(style=style, color=color)

def b_thin(c=SILVER_LINE):
    s = side(c)
    return Border(left=s, right=s, top=s, bottom=s)

def b_med(c=NAVY_MID):
    s = side(c, "medium")
    return Border(left=s, right=s, top=s, bottom=s)

def b_bottom(c=ORANGE, style="medium"):
    return Border(bottom=Side(style=style, color=c))

# ── CELL WRITER ───────────────────────────────────────────────────────────────
def W(ws, row, col, value=None, bold=False, sz=10, color=DARK_TEXT,
      bg=None, ha="left", va="center", wrap=False, fmt=None,
      bdr=None, italic=False, ind=0):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    c.font = Font(name="Arial", size=sz, bold=bold, color=color, italic=italic)
    c.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap, indent=ind)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if fmt:
        c.number_format = fmt
    if bdr:
        c.border = bdr
    return c

def fill(ws, row, c1, c2, bg, bdr=None):
    for col in range(c1, c2 + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=bg)
        if bdr:
            ws.cell(row=row, column=col).border = bdr

def mg(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

def set_row_h(ws, r, h):
    ws.row_dimensions[r].height = h

# ── PHYSICS ───────────────────────────────────────────────────────────────────
def upright_sheet_size(width, depth):
    return width + depth + width + 15

def pipe_beam_sheet_size(height, width):
    return 2 * (height + width)

def roll_form_beam_sheet_size(height, width):
    return ((4 * width) + (2 * height)) + 20

def deep_bar_size(rack_depth):
    return rack_depth - 65

def cross_bar_length(deep_bar, gap, upright_depth):
    a = deep_bar - 50
    b = gap
    return ((a ** 2 + b ** 2) ** 0.5) + 50

def weight(length, width, thickness):
    return (length * width * thickness * 7.85) / 1_000_000

def calc_components(rack):
    uw, ud   = rack["uw"], rack["ud"]
    ul, ut   = rack["ul"], rack["ut"]
    bh, bw   = rack["bh"], rack["bw"]
    bl, bth  = rack["bl"], rack["bth"]
    depth    = rack["depth"]
    dth      = rack["dth"]
    gap      = rack["gap"]
    cth      = rack["cth"]
    levels   = rack["levels"]
    method   = rack["method"]
    main_qty = rack["main_qty"]
    addon_qty= rack["addon_qty"]

    uwid  = upright_sheet_size(uw, ud)
    u_wt  = weight(ul, uwid, ut)
    u_main_qty  = 4 * main_qty
    u_addon_qty = 2 * addon_qty
    u_main  = u_wt * u_main_qty
    u_addon = u_wt * u_addon_qty

    if rack["bt"] == "Pipe Beam":
        bwid = pipe_beam_sheet_size(bh, bw)
    else:
        bwid = roll_form_beam_sheet_size(bh, bw)
    b_wt = weight(bl, bwid, bth) + 1.15

    beam_per_rack = 2 * levels
    b_main_qty  = beam_per_rack * main_qty
    b_addon_qty = beam_per_rack * addon_qty
    b_main  = b_wt * b_main_qty
    b_addon = b_wt * b_addon_qty

    dlen  = deep_bar_size(depth)
    d_wt  = weight(dlen, 92, dth)
    d_main_qty  = 4 * main_qty
    d_addon_qty = 2 * addon_qty
    d_main  = d_wt * d_main_qty
    d_addon = d_wt * d_addon_qty

    eff_h     = ul - method
    num_cross = int(eff_h // gap)
    clen      = cross_bar_length(dlen, gap, ud)
    c_wt      = weight(clen, 92, cth)
    c_per_main  = num_cross * 2
    c_per_addon = num_cross
    c_main_qty  = c_per_main  * main_qty
    c_addon_qty = c_per_addon * addon_qty
    c_main  = c_wt * c_main_qty
    c_addon = c_wt * c_addon_qty

    total_main  = round(u_main  + b_main  + d_main  + c_main,  3)
    total_addon = round(u_addon + b_addon + d_addon + c_addon, 3)

    area_m2        = (bl / 1000) * (depth / 1000)
    udl_kg_m2      = 500
    load_per_level = round(udl_kg_m2 * area_m2, 1)

    return {
        "uwid": round(uwid, 1), "ul": ul, "ut": ut,
        "u_wt": round(u_wt, 4),
        "u_main_qty": u_main_qty, "u_addon_qty": u_addon_qty,
        "u_main": round(u_main, 3), "u_addon": round(u_addon, 3),
        "bwid": round(bwid, 1), "bl": bl, "bth": bth,
        "b_wt": round(b_wt, 4),
        "beam_per_rack": beam_per_rack,
        "b_main_qty": b_main_qty, "b_addon_qty": b_addon_qty,
        "b_main": round(b_main, 3), "b_addon": round(b_addon, 3),
        "dlen": round(dlen, 1), "dth": dth,
        "d_wt": round(d_wt, 4),
        "d_main_qty": d_main_qty, "d_addon_qty": d_addon_qty,
        "d_main": round(d_main, 3), "d_addon": round(d_addon, 3),
        "clen": round(clen, 1), "cth": cth, "num_cross": num_cross,
        "c_per_main": c_per_main, "c_per_addon": c_per_addon,
        "c_wt": round(c_wt, 4),
        "c_main_qty": c_main_qty, "c_addon_qty": c_addon_qty,
        "c_main": round(c_main, 3), "c_addon": round(c_addon, 3),
        "total_main": total_main, "total_addon": total_addon,
        "area_m2": round(area_m2, 3),
        "udl_kg_m2": udl_kg_m2,
        "load_per_level": load_per_level,
    }


def calc_accessories(acc_data, rack_data):
    items = []
    total_uprights = sum(4 * r["main_qty"] + 2 * r["addon_qty"] for r in rack_data)

    cg_qty = acc_data.get("cg_qty", 0)
    cg_wt  = 3.75
    items.append({"name": "Column Guard", "qty": cg_qty, "wt_each": cg_wt,
                  "total_wt": round(cg_qty * cg_wt, 2)})

    rc_qty = acc_data.get("rc_qty", 0)
    rc_wt  = 1.0
    items.append({"name": "Row Connector", "qty": rc_qty, "wt_each": rc_wt,
                  "total_wt": round(rc_qty * rc_wt, 2)})

    for idx, rg in enumerate(acc_data.get("row_guards", []), 1):
        l, qty = rg["l"], rg["qty"]
        h = 400
        wt = (((240 * h * 2 * 7.85) + (240 * l * 2 * 7.85)) * 2) / 1_000_000
        name = f"Row Guard (L:{int(l)} mm)"
        items.append({"name": name, "qty": qty, "wt_each": round(wt, 4),
                      "total_wt": round(wt * qty, 2)})

    for idx, tb in enumerate(acc_data.get("tie_beams", []), 1):
        qty, w, d, l, t = tb["qty"], tb["w"], tb["d"], tb["l"], tb["t"]
        wt = weight(l, upright_sheet_size(w, d), t)
        name = f"Tie Beam (L:{int(l)} mm)"
        items.append({"name": name, "qty": qty, "wt_each": round(wt, 4),
                      "total_wt": round(wt * qty, 2)})

    for idx, bps in enumerate(acc_data.get("bps_list", []), 1):
        qty, l = bps["qty"], bps["l"]
        wt = ((160 * 1.6 * l * 7.85) / 1_000_000) + 0.6
        name = f"Back Pallet Stopper (L:{int(l)} mm)"
        items.append({"name": name, "qty": qty, "wt_each": round(wt, 4),
                      "total_wt": round(wt * qty, 2)})

    items.append({"name": "Base Plate", "qty": total_uprights,
                  "wt_each": "Incl.", "total_wt": "-"})

    return items


def calc_accessories_quotation(acc_data, rack_data):
    items = []

    cg_qty = acc_data.get("cg_qty", 0)
    if cg_qty > 0:
        cg_wt = 3.75
        items.append({"name": "Column Guard", "qty": cg_qty, "wt_each": cg_wt,
                      "total_wt": round(cg_qty * cg_wt, 2)})

    rc_qty = acc_data.get("rc_qty", 0)
    if rc_qty > 0:
        rc_wt = 1.0
        items.append({"name": "Row Connector", "qty": rc_qty, "wt_each": rc_wt,
                      "total_wt": round(rc_qty * rc_wt, 2)})

    for idx, rg in enumerate(acc_data.get("row_guards", []), 1):
        l, qty = rg["l"], rg["qty"]
        if qty > 0:
            h = 400
            wt = (((240 * h * 2 * 7.85) + (240 * l * 2 * 7.85)) * 2) / 1_000_000
            name = f"Row Guard (L:{int(l)} mm)"
            items.append({"name": name, "qty": qty, "wt_each": round(wt, 4),
                          "total_wt": round(wt * qty, 2)})

    for idx, tb in enumerate(acc_data.get("tie_beams", []), 1):
        qty, w, d, l, t = tb["qty"], tb["w"], tb["d"], tb["l"], tb["t"]
        if qty > 0:
            wt = weight(l, upright_sheet_size(w, d), t)
            name = f"Tie Beam (L:{int(l)} mm)"
            items.append({"name": name, "qty": qty, "wt_each": round(wt, 4),
                          "total_wt": round(wt * qty, 2)})

    for idx, bps in enumerate(acc_data.get("bps_list", []), 1):
        qty, l = bps["qty"], bps["l"]
        if qty > 0:
            wt = ((160 * 1.6 * l * 7.85) / 1_000_000) + 0.6
            name = f"Back Pallet Stopper (L:{int(l)} mm)"
            items.append({"name": name, "qty": qty, "wt_each": round(wt, 4),
                          "total_wt": round(wt * qty, 2)})

    return items


# ═══════════════════════════════════════════════════════════════════════════════
#  SHEET 1 — COMMERCIAL OFFER
# ═══════════════════════════════════════════════════════════════════════════════
def build_quotation_sheet(ws, client, product, offer_no, date_obj,
                          project_name, rack_data, rate_per_kg,
                          acc_data=None, logo_path=None):
    ws.sheet_view.showGridLines = False

    col_w = {
        1: 1.2, 2: 5, 3: 6, 4: 34, 5: 13,
        6: 10, 7: 10, 8: 10, 9: 10, 10: 17, 11: 18, 12: 1.2
    }
    for col, w in col_w.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    R = 1

    # ── Top orange accent bar ─────────────────────────────────────────────────
    set_row_h(ws, R, 6); fill(ws, R, 1, 12, ORANGE); R += 1

    # ── Logo + Company Header ─────────────────────────────────────────────────
    set_row_h(ws, R, 26); fill(ws, R, 1, 12, NAVY)
    if logo_path and os.path.exists(logo_path):
        try:
            img = XLImage(logo_path)
            img.width = 160; img.height = 50
            img.anchor = f"B{R}"
            ws.add_image(img)
        except Exception:
            pass
    mg(ws, R, 3, R, 11)
    c = ws.cell(row=R, column=3)
    c.value = "BRIJ INDUSTRIES"
    c.font = Font(name="Arial", size=15, bold=True, color=WHITE)
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.fill = PatternFill("solid", fgColor=NAVY)
    R += 1

    # ── DSS Brand row — BIG, PROMINENT ───────────────────────────────────────
    set_row_h(ws, R, 52); fill(ws, R, 1, 12, NAVY)
    mg(ws, R, 2, R, 5)
    c_dss = ws.cell(row=R, column=2)
    c_dss.value = "DSS"
    c_dss.font = Font(name="Arial", size=38, bold=True, color=ORANGE)
    c_dss.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c_dss.fill = PatternFill("solid", fgColor=NAVY)
    mg(ws, R, 6, R, 11)
    c_full = ws.cell(row=R, column=6)
    c_full.value = "DOLPHIN  STORAGE  SOLUTIONS"
    c_full.font = Font(name="Arial", size=13, bold=True, color=WHITE)
    c_full.alignment = Alignment(horizontal="left", vertical="center")
    c_full.fill = PatternFill("solid", fgColor=NAVY)
    R += 1

    # ── Contact strip ─────────────────────────────────────────────────────────
    set_row_h(ws, R, 14); fill(ws, R, 1, 12, NAVY_MID)
    mg(ws, R, 2, R, 11)
    W(ws, R, 2,
      "86/3/1 Road No 7, Mundka Industrial Area, New Delhi – 110041   |   GST: 07AAMFB6403G1ZM   |   +91 9625589161 / 9811096149   |   brijindustries09@rediffmail.com",
      sz=7.5, color="A8CCF0", bg=NAVY_MID, ha="center", italic=True)
    R += 1

    # ── Orange divider ────────────────────────────────────────────────────────
    set_row_h(ws, R, 5); fill(ws, R, 1, 12, ORANGE); R += 1
    set_row_h(ws, R, 7); R += 1

    # ── COMMERCIAL OFFER title ────────────────────────────────────────────────
    set_row_h(ws, R, 28)
    mg(ws, R, 2, R, 11)
    c = ws.cell(row=R, column=2)
    c.value = "C O M M E R C I A L   O F F E R"
    c.font = Font(name="Arial", size=13, bold=True, color=WHITE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.border = b_med(ORANGE)
    R += 1

    set_row_h(ws, R, 5); fill(ws, R, 2, 11, ORANGE); R += 1
    set_row_h(ws, R, 7); R += 1

    # ── Detail pairs ──────────────────────────────────────────────────────────
    def detail_pair(lbl_l, val_l, lbl_r, val_r, bg=ROW_ALT):
        nonlocal R
        set_row_h(ws, R, 22)
        fill(ws, R, 2, 11, bg, b_thin())
        mg(ws, R, 2, R, 3)
        W(ws, R, 2, lbl_l, bold=True, sz=9, color=NAVY_MID, bg=bg, ha="right", bdr=b_thin())
        mg(ws, R, 4, R, 6)
        W(ws, R, 4, val_l, sz=9, color=DARK_TEXT, bg=bg, ind=1, bdr=b_thin())
        mg(ws, R, 7, R, 8)
        W(ws, R, 7, lbl_r, bold=True, sz=9, color=NAVY_MID, bg=bg, ha="right", bdr=b_thin())
        mg(ws, R, 9, R, 11)
        W(ws, R, 9, val_r, sz=9, color=DARK_TEXT, bg=bg, ind=1, bdr=b_thin())
        R += 1

    detail_pair("To :",      client,       "Date :",      date_obj.strftime("%d %B %Y"), WHITE)
    detail_pair("Product :", product,      "Offer No. :", offer_no,                      ROW_ALT)
    detail_pair("Project :", project_name, "",            "",                             WHITE)

    set_row_h(ws, R, 9); R += 1

    # ── TECHNICAL DETAILS ─────────────────────────────────────────────────────
    set_row_h(ws, R, 24)
    mg(ws, R, 2, R, 11)
    c = ws.cell(row=R, column=2)
    c.value = "TECHNICAL DETAILS"
    c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor=NAVY_MID)
    c.border = b_med(ORANGE)
    R += 1

    set_row_h(ws, R, 4); fill(ws, R, 2, 11, ORANGE); R += 1

    # ── Technical headers — Beam Type & Upright Section REMOVED ──────────────
    set_row_h(ws, R, 30); fill(ws, R, 2, 11, NAVY)
    tech_hdrs = [
        (2,  "MODULE",             "center"),
        (3,  "",                   "center"),   # merged with MODULE
        (4,  "HEIGHT\n(mm)",       "center"),
        (5,  "LENGTH\n(mm)",       "center"),
        (6,  "DEPTH\n(mm)",        "center"),
        (7,  "LEVELS",             "center"),
        (8,  "UDL\n(kg/m²)",       "center"),
        (9,  "LOAD /\nLEVEL (kg)", "center"),
        (10, "",                   "center"),   # merged with LOAD
        (11, "PALLETS\n/ LEVEL",   "center"),
    ]
    for col, txt, al in tech_hdrs:
        c = ws.cell(row=R, column=col)
        c.value = txt
        c.font = Font(name="Arial", size=8.5, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal=al, vertical="center", wrap_text=True)
        c.border = b_thin(NAVY_MID)
    # Merge MODULE and LOAD headers for cleaner look
    mg(ws, R, 2, R, 3)
    mg(ws, R, 9, R, 10)
    R += 1

    for idx, rack in enumerate(rack_data):
        comp = calc_components(rack)
        bg = WHITE if idx % 2 == 0 else ROW_ALT
        set_row_h(ws, R, 21)
        fill(ws, R, 2, 11, bg, b_thin())

        mg(ws, R, 2, R, 3)
        W(ws, R, 2, rack['module'], bold=True, sz=9, color=NAVY_MID, bg=bg, ha="center", bdr=b_thin())
        W(ws, R, 4, rack["ul"],     sz=9, bg=bg, ha="center", bdr=b_thin())
        W(ws, R, 5, rack["bl"],     sz=9, bg=bg, ha="center", bdr=b_thin())
        W(ws, R, 6, rack["depth"],  sz=9, bg=bg, ha="center", bdr=b_thin())
        W(ws, R, 7, rack["levels"], sz=9, bg=bg, ha="center", bdr=b_thin())
        W(ws, R, 8, comp["udl_kg_m2"], sz=9, bg=bg, ha="center", bdr=b_thin())
        mg(ws, R, 9, R, 10)
        W(ws, R, 9, comp["load_per_level"], sz=9, bg=bg, ha="center", bdr=b_thin(), fmt="#,##0")
        pallets = max(1, int(rack["bl"] / 1200))
        W(ws, R, 11, pallets, sz=9, bg=bg, ha="center", bdr=b_thin())
        R += 1

    set_row_h(ws, R, 4); fill(ws, R, 2, 11, ORANGE); R += 1
    set_row_h(ws, R, 9); R += 1

    # ── SCOPE OF SUPPLY ───────────────────────────────────────────────────────
    set_row_h(ws, R, 24)
    mg(ws, R, 2, R, 11)
    c = ws.cell(row=R, column=2)
    c.value = "SCOPE OF SUPPLY"
    c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor=NAVY_MID)
    c.border = b_med(ORANGE)
    R += 1

    set_row_h(ws, R, 4); fill(ws, R, 2, 11, ORANGE); R += 1

    set_row_h(ws, R, 28); fill(ws, R, 2, 11, NAVY)
    scope_hdrs = [
        (2,  "SR.",              "center"),
        (3,  "",                 "center"),
        (4,  "DESCRIPTION",      "left"),
        (5,  "TYPE",             "center"),
        (6,  "",                 "center"),
        (7,  "",                 "center"),
        (8,  "QTY",              "center"),
        (9,  "",                 "center"),
        (10, "UNIT PRICE (Rs.)", "right"),
        (11, "AMOUNT (Rs.)",     "right"),
    ]
    for col, txt, al in scope_hdrs:
        c = ws.cell(row=R, column=col)
        c.value = txt
        c.font = Font(name="Arial", size=8.5, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal=al, vertical="center")
        c.border = b_thin(NAVY_MID)
    R += 1

    total_basic = 0.0
    sr = 1

    for rack in rack_data:
        comp = calc_components(rack)
        main_price  = comp["total_main"]  * rate_per_kg
        addon_price = comp["total_addon"] * rate_per_kg
        main_total  = main_price * rack["main_qty"]
        total_basic += main_total

        bg = WHITE if sr % 2 == 1 else ROW_ALT
        set_row_h(ws, R, 21); fill(ws, R, 2, 11, bg, b_thin())
        W(ws, R, 2, sr, sz=9, color=MID_TEXT, bg=bg, ha="center", bdr=b_thin())
        mg(ws, R, 3, R, 6)
        W(ws, R, 3, f"  {rack['module']}  Main Rack", bold=True, sz=9, color=DARK_TEXT, bg=bg, bdr=b_thin())
        W(ws, R, 7, "Main", sz=8, color=NAVY_MID, bg=bg, ha="center", bdr=b_thin())
        mg(ws, R, 8, R, 9)
        W(ws, R, 8, rack["main_qty"], sz=9, bg=bg, ha="center", bdr=b_thin())
        W(ws, R, 10, round(main_price, 2), sz=9, bg=bg, ha="right", fmt="#,##0.00", bdr=b_thin())
        W(ws, R, 11, round(main_total, 2), bold=True, sz=9, color=NAVY, bg=bg, ha="right", fmt="#,##0.00", bdr=b_thin())
        R += 1; sr += 1

        if rack["addon_qty"] > 0:
            addon_total  = addon_price * rack["addon_qty"]
            total_basic += addon_total
            bg = WHITE if sr % 2 == 1 else ROW_ALT
            set_row_h(ws, R, 21); fill(ws, R, 2, 11, bg, b_thin())
            W(ws, R, 2, sr, sz=9, color=MID_TEXT, bg=bg, ha="center", bdr=b_thin())
            mg(ws, R, 3, R, 6)
            W(ws, R, 3, f"  {rack['module']}  Add-on Rack", sz=9, color=MID_TEXT, bg=bg, bdr=b_thin())
            W(ws, R, 7, "Add-on", sz=8, color=ORANGE_DARK, bg=bg, ha="center", bdr=b_thin())
            mg(ws, R, 8, R, 9)
            W(ws, R, 8, rack["addon_qty"], sz=9, bg=bg, ha="center", bdr=b_thin())
            W(ws, R, 10, round(addon_price, 2), sz=9, bg=bg, ha="right", fmt="#,##0.00", bdr=b_thin())
            W(ws, R, 11, round(addon_total, 2), bold=True, sz=9, color=NAVY, bg=bg, ha="right", fmt="#,##0.00", bdr=b_thin())
            R += 1; sr += 1

    if acc_data:
        acc_items = calc_accessories_quotation(acc_data, rack_data)
        if acc_items:
            set_row_h(ws, R, 18); fill(ws, R, 2, 11, TABLE_DARK)
            mg(ws, R, 2, R, 11)
            c = ws.cell(row=R, column=2)
            c.value = "  ACCESSORIES"
            c.font = Font(name="Arial", size=8.5, bold=True, color=NAVY_MID)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.fill = PatternFill("solid", fgColor=TABLE_DARK)
            c.border = b_thin()
            R += 1

            for acc in acc_items:
                bg = ORANGE_LIGHT if sr % 2 == 1 else WHITE
                set_row_h(ws, R, 19); fill(ws, R, 2, 11, bg, b_thin())
                W(ws, R, 2, sr, sz=9, color=MID_TEXT, bg=bg, ha="center", bdr=b_thin())
                mg(ws, R, 3, R, 7)
                W(ws, R, 3, f"  {acc['name']}", sz=9, color=DARK_TEXT, bg=bg, bdr=b_thin())
                mg(ws, R, 8, R, 9)
                W(ws, R, 8, acc["qty"], sz=9, bg=bg, ha="center", bdr=b_thin())
                wt_each = acc["wt_each"]
                if isinstance(wt_each, (int, float)):
                    W(ws, R, 10, round(wt_each, 3), sz=9, bg=bg, ha="right", fmt="#,##0.000", bdr=b_thin())
                else:
                    W(ws, R, 10, str(wt_each), sz=9, bg=bg, ha="center", bdr=b_thin())
                tot = acc["total_wt"]
                if isinstance(tot, (int, float)):
                    W(ws, R, 11, f"{tot:.2f} kg", sz=9, color=NAVY_MID, bg=bg, ha="right", bdr=b_thin())
                else:
                    W(ws, R, 11, str(tot), sz=9, color=NAVY_MID, bg=bg, ha="center", bdr=b_thin())
                R += 1; sr += 1

    set_row_h(ws, R, 4); fill(ws, R, 2, 11, ORANGE); R += 1

    # ── Subtotal ──────────────────────────────────────────────────────────────
    set_row_h(ws, R, 26); fill(ws, R, 2, 11, NAVY_MID)
    mg(ws, R, 2, R, 10)
    c = ws.cell(row=R, column=2)
    c.value = "SUBTOTAL  BASIC AMOUNT  (Excl. GST)"
    c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c.fill = PatternFill("solid", fgColor=NAVY_MID)
    c.border = b_med(NAVY_MID)
    W(ws, R, 11, round(total_basic, 2), bold=True, sz=10, color=WHITE,
      bg=NAVY, ha="right", fmt="#,##0.00", bdr=b_med(NAVY))
    R += 1
    set_row_h(ws, R, 9); R += 1

    gst   = round(total_basic * 0.18, 2)
    grand = round(total_basic + gst, 2)

    def price_row(lbl, val, bg=WHITE, bold=False, vc=DARK_TEXT, is_text=False):
        nonlocal R
        set_row_h(ws, R, 22); fill(ws, R, 7, 11, bg)
        mg(ws, R, 7, R, 10)
        c_l = ws.cell(row=R, column=7)
        c_l.value = lbl
        c_l.font = Font(name="Arial", size=9.5, bold=bold, color=vc)
        c_l.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c_l.fill = PatternFill("solid", fgColor=bg)
        c_l.border = b_thin()
        c_v = ws.cell(row=R, column=11)
        c_v.value = val
        c_v.font = Font(name="Arial", size=9.5, bold=bold, color=vc, italic=is_text)
        c_v.alignment = Alignment(horizontal="center" if is_text else "right", vertical="center", indent=1)
        c_v.fill = PatternFill("solid", fgColor=bg)
        c_v.border = b_thin()
        if not is_text:
            c_v.number_format = "#,##0.00"
        R += 1

    price_row("Basic Amount (Rs.)",  round(total_basic, 2), ROW_ALT,      True,  NAVY)
    price_row("Freight Charges",     "Inclusive",           ORANGE_LIGHT, False, ORANGE_DARK, True)
    price_row("Erection Charges",    "Inclusive",           ORANGE_LIGHT, False, ORANGE_DARK, True)
    price_row("GST @ 18%  (Rs.)",    gst,                   WHITE,        False, MID_TEXT)

    # ── Spacer before grand total ──────────────────────────────────────────────
    set_row_h(ws, R, 10); R += 1

    # ── GRAND TOTAL — isolated row, no merging issue ──────────────────────────
    set_row_h(ws, R, 36); fill(ws, R, 2, 11, NAVY)
    mg(ws, R, 2, R, 10)
    c = ws.cell(row=R, column=2)
    c.value = "GRAND TOTAL  ·  Inclusive of GST @ 18%"
    c.font = Font(name="Arial", size=12, bold=True, color=WHITE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor=NAVY)
    # Explicit border only on the merged region sides
    c.border = Border(
        left=Side(style="medium", color=ORANGE),
        top=Side(style="medium", color=ORANGE),
        bottom=Side(style="medium", color=ORANGE),
    )
    c_gv = ws.cell(row=R, column=11)
    c_gv.value = grand
    c_gv.font = Font(name="Arial", size=12, bold=True, color=WHITE)
    c_gv.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c_gv.fill = PatternFill("solid", fgColor=ORANGE)
    c_gv.border = Border(
        right=Side(style="medium", color=ORANGE),
        top=Side(style="medium", color=ORANGE),
        bottom=Side(style="medium", color=ORANGE),
    )
    c_gv.number_format = 'Rs. #,##0.00'
    R += 1

    # ── Spacer after grand total (prevents visual merge with next row) ─────────
    set_row_h(ws, R, 8); R += 1

    # ── Orange accent bar ─────────────────────────────────────────────────────
    set_row_h(ws, R, 5); fill(ws, R, 1, 12, ORANGE); R += 1
    set_row_h(ws, R, 9); R += 1

    # ── Terms & Bank ──────────────────────────────────────────────────────────
    set_row_h(ws, R, 22); fill(ws, R, 2, 11, NAVY_MID)
    mg(ws, R, 2, R, 6)
    c = ws.cell(row=R, column=2)
    c.value = "TERMS & CONDITIONS"
    c.font = Font(name="Arial", size=9, bold=True, color=WHITE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor=NAVY_MID)
    mg(ws, R, 7, R, 11)
    c2 = ws.cell(row=R, column=7)
    c2.value = "BANK DETAILS"
    c2.font = Font(name="Arial", size=9, bold=True, color=WHITE)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c2.fill = PatternFill("solid", fgColor=NAVY_MID)
    R += 1

    terms = [
        "Payment: 50% advance, balance against delivery",
        "Delivery: 10 to 12 weeks from advance receipt",
        "Warranty: 12 months from commissioning date",
        "GST @ 18% applicable as per Government norms",
        "Prices valid for 4 to 5 days from date of offer",
    ]
    bank = [
        "Account Name :  BRIJ INDUSTRIES",
        "Bank              :  ICICI Bank Ltd.",
        "Account No.   :  135805001108",
        "IFSC Code      :  ICIC0001358",
        "Branch            :  Mundka, New Delhi",
    ]
    for t, b in zip(terms, bank):
        set_row_h(ws, R, 18)
        mg(ws, R, 2, R, 6)
        c = ws.cell(row=R, column=2)
        c.value = f"  ›  {t}"
        c.font = Font(name="Arial", size=8.5, color=MID_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.fill = PatternFill("solid", fgColor=ROW_ALT)
        c.border = b_thin()
        mg(ws, R, 7, R, 11)
        c2 = ws.cell(row=R, column=7)
        c2.value = f"  {b}"
        c2.font = Font(name="Arial", size=8.5, color=MID_TEXT)
        c2.alignment = Alignment(horizontal="left", vertical="center")
        c2.fill = PatternFill("solid", fgColor=ORANGE_LIGHT)
        c2.border = b_thin()
        R += 1

    set_row_h(ws, R, 9); R += 1

    set_row_h(ws, R, 44)
    mg(ws, R, 2, R, 6)
    c = ws.cell(row=R, column=2)
    c.value = "Customer Signature & Stamp"
    c.font = Font(name="Arial", size=8, color=LIGHT_TEXT, italic=True)
    c.alignment = Alignment(horizontal="center", vertical="bottom")
    c.border = Border(top=Side(style="medium", color=NAVY_MID))
    mg(ws, R, 7, R, 11)
    c2 = ws.cell(row=R, column=7)
    c2.value = "For BRIJ INDUSTRIES  |  DSS Dolphin Storage Solutions"
    c2.font = Font(name="Arial", size=8, color=LIGHT_TEXT, italic=True)
    c2.alignment = Alignment(horizontal="center", vertical="bottom")
    c2.border = Border(top=Side(style="medium", color=NAVY_MID))
    R += 1

    set_row_h(ws, R, 6); fill(ws, R, 1, 12, ORANGE); R += 1
    set_row_h(ws, R, 16); fill(ws, R, 1, 12, NAVY)
    mg(ws, R, 2, R, 11)
    W(ws, R, 2, "DSS  ·  Dolphin Storage Solutions  ·  BRIJ INDUSTRIES  ·  www.brijindustries.in",
      sz=8.5, color="A8CCF0", bg=NAVY, ha="center", italic=True)

    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.print_area = f"A1:{get_column_letter(12)}{R}"

    return total_basic, gst, grand


# ═══════════════════════════════════════════════════════════════════════════════
#  SHEET 2 — BILL OF MATERIALS
# ═══════════════════════════════════════════════════════════════════════════════
def build_bom_sheet(ws, client, offer_no, date_obj, rack_data, acc_data=None):
    ws.sheet_view.showGridLines = False

    col_w = {
        1: 1.2, 2: 4, 3: 26, 4: 13, 5: 11,
        6: 11, 7: 12, 8: 11, 9: 11, 10: 15, 11: 15, 12: 16, 13: 1.2
    }
    for col, w in col_w.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    R = 1
    set_row_h(ws, R, 6); fill(ws, R, 1, 13, ORANGE); R += 1

    # DSS as primary brand in BOM header
    set_row_h(ws, R, 44); fill(ws, R, 1, 13, NAVY)
    mg(ws, R, 2, R, 5)
    c = ws.cell(row=R, column=2)
    c.value = "DSS"
    c.font = Font(name="Arial", size=30, bold=True, color=ORANGE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.fill = PatternFill("solid", fgColor=NAVY)
    mg(ws, R, 6, R, 12)
    c2 = ws.cell(row=R, column=6)
    c2.value = "DOLPHIN STORAGE SOLUTIONS  —  BILL OF MATERIALS"
    c2.font = Font(name="Arial", size=14, bold=True, color=WHITE)
    c2.alignment = Alignment(horizontal="left", vertical="center")
    c2.fill = PatternFill("solid", fgColor=NAVY)
    R += 1

    set_row_h(ws, R, 14); fill(ws, R, 1, 13, NAVY_MID)
    mg(ws, R, 2, R, 12)
    W(ws, R, 2,
      f"BRIJ INDUSTRIES   |   Offer No: {offer_no}   |   Customer: {client}   |   Date: {date_obj.strftime('%d %B %Y')}",
      sz=9, color="A8CCF0", bg=NAVY_MID, ha="center")
    R += 1

    set_row_h(ws, R, 5); fill(ws, R, 1, 13, ORANGE); R += 1
    set_row_h(ws, R, 7); R += 1

    grand_main_wt  = 0.0
    grand_addon_wt = 0.0

    for rack in rack_data:
        comp = calc_components(rack)

        set_row_h(ws, R, 23)
        mg(ws, R, 2, R, 12)
        c = ws.cell(row=R, column=2)
        c.value = (f"  {rack['module']}   |   Main Racks: {rack['main_qty']}"
                   f"   |   Add-on Racks: {rack['addon_qty']}   |   Levels: {rack['levels']}"
                   f"   |   Cross Method: {rack['method']} mm   |   Gap: {rack['gap']} mm")
        c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.fill = PatternFill("solid", fgColor=NAVY_MID)
        c.border = b_med(ORANGE)
        R += 1

        set_row_h(ws, R, 4); fill(ws, R, 2, 12, ORANGE); R += 1

        set_row_h(ws, R, 34); fill(ws, R, 2, 12, NAVY)
        BOM_H = [
            (2,  "SR.",                  "center"),
            (3,  "COMPONENT",            "left"),
            (4,  "SECTION\nPROFILE",     "center"),
            (5,  "LENGTH\n(mm)",         "center"),
            (6,  "THICK\n(mm)",          "center"),
            (7,  "WT / PCS\n(kg)",       "center"),
            (8,  "QTY /\nMAIN",          "center"),
            (9,  "QTY /\nADD-ON",        "center"),
            (10, "MAIN\nTOTAL (kg)",     "center"),
            (11, "ADD-ON\nTOTAL (kg)",   "center"),
            (12, "LOAD /\nLEVEL (kg)",   "center"),
        ]
        for col, txt, al in BOM_H:
            c = ws.cell(row=R, column=col)
            c.value = txt
            c.font = Font(name="Arial", size=8, bold=True, color=WHITE)
            c.fill = PatternFill("solid", fgColor=NAVY)
            c.alignment = Alignment(horizontal=al, vertical="center", wrap_text=True)
            c.border = b_thin(NAVY_MID)
        R += 1

        def comp_row(sr_n, comp_name, section, length, thick, wt_each,
                     qty_main, qty_addon, load_level_val, bg):
            nonlocal R
            set_row_h(ws, R, 20)
            fill(ws, R, 2, 12, bg, b_thin())
            main_tot  = round(wt_each * qty_main,  3)
            addon_tot = round(wt_each * qty_addon, 3)
            row_vals = [
                (2,  sr_n,             "center", None),
                (3,  f"  {comp_name}", "left",   None),
                (4,  section,          "center", None),
                (5,  length,           "center", None),
                (6,  thick,            "center", None),
                (7,  wt_each,          "center", "#,##0.000"),
                (8,  qty_main,         "center", None),
                (9,  qty_addon,        "center", None),
                (10, main_tot,         "right",  "#,##0.000"),
                (11, addon_tot,        "right",  "#,##0.000"),
                (12, load_level_val,   "center", "#,##0"),
            ]
            for col, val, al, nf in row_vals:
                c = ws.cell(row=R, column=col)
                c.value = val
                c.font = Font(name="Arial", size=8.5, color=DARK_TEXT)
                c.alignment = Alignment(horizontal=al, vertical="center")
                c.fill = PatternFill("solid", fgColor=bg)
                c.border = b_thin()
                if nf:
                    c.number_format = nf
            R += 1

        if rack["bt"] == "Pipe Beam":
            beam_sec = f"{rack['bh']}x{rack['bw']} Pipe"
        else:
            beam_sec = f"{rack['bh']}x{rack['bw']} Roll"

        comp_row(1, "Upright / Column",
                 f"{rack['uw']}x{rack['ud']} Box",
                 rack["ul"], rack["ut"], comp["u_wt"],
                 comp["u_main_qty"], comp["u_addon_qty"], "—", WHITE)

        comp_row(2, f"Beam  [2 x {rack['levels']} levels = {comp['beam_per_rack']} per rack]",
                 beam_sec, rack["bl"], rack["bth"], comp["b_wt"],
                 comp["b_main_qty"], comp["b_addon_qty"], comp["load_per_level"], ROW_ALT)

        comp_row(3, "Deep Bar (Shelf Support)",
                 "92 mm flat", int(comp["dlen"]), rack["dth"], comp["d_wt"],
                 comp["d_main_qty"], comp["d_addon_qty"], "—", WHITE)

        comp_row(4, f"Cross Brace  [{comp['num_cross']} crosses x 2 per main, x 1 per addon]",
                 "92 mm flat", int(comp["clen"]), rack["cth"], comp["c_wt"],
                 comp["c_main_qty"], comp["c_addon_qty"], "—", ROW_ALT)

        total_m_wt = round(comp["total_main"]  * rack["main_qty"],  2)
        total_a_wt = round(comp["total_addon"] * rack["addon_qty"], 2)
        grand_main_wt  += total_m_wt
        grand_addon_wt += total_a_wt

        def mod_summary(lbl, v_main, v_addon, bg, bold=False, vc=DARK_TEXT, nf="#,##0.00"):
            nonlocal R
            set_row_h(ws, R, 21); fill(ws, R, 2, 12, bg)
            mg(ws, R, 2, R, 9)
            c = ws.cell(row=R, column=2)
            c.value = lbl
            c.font = Font(name="Arial", size=9, bold=bold, color=vc)
            c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
            c.fill = PatternFill("solid", fgColor=bg)
            c.border = b_med(NAVY_MID) if bold else b_thin()
            for col, val in [(10, v_main), (11, v_addon)]:
                cv = ws.cell(row=R, column=col)
                cv.value = val
                cv.font = Font(name="Arial", size=9, bold=bold, color=vc)
                cv.alignment = Alignment(horizontal="right", vertical="center", indent=1)
                cv.fill = PatternFill("solid", fgColor=bg)
                cv.border = b_med(NAVY_MID) if bold else b_thin()
                cv.number_format = nf
            R += 1

        mod_summary("Wt  Single Rack (kg)", comp["total_main"], comp["total_addon"], TABLE_DARK)
        mod_summary(f"Total Wt  All Racks  [Main x{rack['main_qty']}  |  Add-on x{rack['addon_qty']}]",
                    total_m_wt, total_a_wt, TABLE_DARK, bold=True, vc=NAVY)

        set_row_h(ws, R, 4); fill(ws, R, 2, 12, ORANGE); R += 1
        set_row_h(ws, R, 10); R += 1

    if acc_data:
        acc_items = calc_accessories(acc_data, rack_data)
        set_row_h(ws, R, 22)
        mg(ws, R, 2, R, 12)
        c = ws.cell(row=R, column=2)
        c.value = "ACCESSORIES"
        c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = PatternFill("solid", fgColor=STEEL_BLUE)
        c.border = b_med(NAVY)
        R += 1

        set_row_h(ws, R, 4); fill(ws, R, 2, 12, ORANGE); R += 1

        set_row_h(ws, R, 28); fill(ws, R, 2, 12, NAVY)
        ACC_H = [
            (2,  "SR.",           "center"),
            (3,  "ITEM",          "left"),
            (5,  "QTY",           "center"),
            (6,  "WT / PCS (kg)", "center"),
            (10, "TOTAL WT (kg)", "center"),
        ]
        for col in range(2, 13):
            c = ws.cell(row=R, column=col)
            c.font = Font(name="Arial", size=8, bold=True, color=WHITE)
            c.fill = PatternFill("solid", fgColor=NAVY)
            c.border = b_thin(NAVY_MID)
        for col, txt, al in ACC_H:
            c = ws.cell(row=R, column=col)
            c.value = txt
            c.alignment = Alignment(horizontal=al, vertical="center", wrap_text=True)
        R += 1

        for idx, acc in enumerate(acc_items):
            bg = BLUE_LIGHT if idx % 2 == 0 else WHITE
            set_row_h(ws, R, 20); fill(ws, R, 2, 12, bg, b_thin())
            W(ws, R, 2, idx + 1, sz=9, color=MID_TEXT, bg=bg, ha="center", bdr=b_thin())
            mg(ws, R, 3, R, 4)
            W(ws, R, 3, f"  {acc['name']}", sz=9, color=NAVY_LIGHT, bg=bg, bold=True, bdr=b_thin())
            W(ws, R, 5, acc["qty"], sz=9, bg=bg, ha="center", bdr=b_thin())
            wt_each = acc["wt_each"]
            if isinstance(wt_each, (int, float)):
                W(ws, R, 6, round(wt_each, 4), sz=9, bg=bg, ha="right", fmt="#,##0.000", bdr=b_thin())
            else:
                W(ws, R, 6, str(wt_each), sz=9, bg=bg, ha="center", bdr=b_thin())
            tot = acc["total_wt"]
            if isinstance(tot, (int, float)):
                W(ws, R, 10, round(tot, 2), sz=9, color=NAVY, bg=bg,
                  ha="right", bold=True, fmt="#,##0.00", bdr=b_thin())
            else:
                W(ws, R, 10, str(tot), sz=9, color=NAVY, bg=bg, ha="center", bdr=b_thin())
            R += 1

        set_row_h(ws, R, 9); R += 1

    set_row_h(ws, R, 5); fill(ws, R, 1, 13, ORANGE); R += 1
    set_row_h(ws, R, 24)
    mg(ws, R, 2, R, 12)
    c = ws.cell(row=R, column=2)
    c.value = "TOTAL TONNAGE SUMMARY"
    c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.border = b_med(ORANGE)
    R += 1

    set_row_h(ws, R, 26); fill(ws, R, 2, 12, NAVY)
    ton_hdrs = [
        (2,  "MODULE",                    "left"),
        (6,  "ALL MAIN RACKS\nWT (kg)",   "center"),
        (8,  "ALL ADD-ON RACKS\nWT (kg)", "center"),
        (10, "COMBINED\nWT (kg)",         "center"),
        (11, "COMBINED\nWT (MT)",         "center"),
        (12, "LOAD /\nLEVEL (kg)",        "center"),
    ]
    for col, txt, al in ton_hdrs:
        c = ws.cell(row=R, column=col)
        c.value = txt
        c.font = Font(name="Arial", size=8, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal=al, vertical="center", wrap_text=True)
        c.border = b_thin(NAVY_MID)
    R += 1

    for idx, rack in enumerate(rack_data):
        comp = calc_components(rack)
        m_wt = round(comp["total_main"]  * rack["main_qty"],  2)
        a_wt = round(comp["total_addon"] * rack["addon_qty"], 2)
        comb = round(m_wt + a_wt, 2)
        bg   = WHITE if idx % 2 == 0 else ROW_ALT

        set_row_h(ws, R, 20); fill(ws, R, 2, 12, bg, b_thin())
        mg(ws, R, 2, R, 5)
        W(ws, R, 2, f"  {rack['module']}  (Main x{rack['main_qty']} | Add-on x{rack['addon_qty']})",
          bold=True, sz=9, color=NAVY_MID, bg=bg, bdr=b_thin())
        W(ws, R, 6,  m_wt,               sz=9, bg=bg, ha="right", fmt="#,##0.00",  bdr=b_thin())
        W(ws, R, 8,  a_wt,               sz=9, bg=bg, ha="right", fmt="#,##0.00",  bdr=b_thin())
        W(ws, R, 10, comb,               sz=9, bg=bg, ha="right", fmt="#,##0.00",  bdr=b_thin())
        W(ws, R, 11, round(comb/1000,3), sz=9, bg=bg, ha="right", fmt="#,##0.000", bdr=b_thin())
        W(ws, R, 12, comp["load_per_level"], sz=9, bg=bg, ha="center", fmt="#,##0", bdr=b_thin())
        R += 1

    grand_comb = round(grand_main_wt + grand_addon_wt, 2)
    set_row_h(ws, R, 24); fill(ws, R, 2, 12, NAVY)
    mg(ws, R, 2, R, 5)
    c = ws.cell(row=R, column=2)
    c.value = "  GRAND TOTAL TONNAGE"
    c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.border = b_med(ORANGE)
    for col, val, nf in [
        (6,  round(grand_main_wt, 2),  "#,##0.00"),
        (8,  round(grand_addon_wt, 2), "#,##0.00"),
        (10, grand_comb,               "#,##0.00"),
        (11, round(grand_comb/1000,3), "#,##0.000"),
    ]:
        cv = ws.cell(row=R, column=col)
        cv.value = val
        cv.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        cv.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        cv.fill = PatternFill("solid", fgColor=NAVY_MID)
        cv.border = b_med(ORANGE)
        cv.number_format = nf
    R += 1

    set_row_h(ws, R, 6); fill(ws, R, 1, 13, ORANGE); R += 1
    set_row_h(ws, R, 16); fill(ws, R, 1, 13, NAVY)
    mg(ws, R, 2, R, 12)
    W(ws, R, 2, "DSS  ·  Dolphin Storage Solutions  ·  BRIJ INDUSTRIES  ·  www.brijindustries.in",
      sz=8.5, color="A8CCF0", bg=NAVY, ha="center", italic=True)

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.print_area = f"A1:{get_column_letter(13)}{R}"


# ═══════════════════════════════════════════════════════════════════════════════
#  MASTER BUILD
# ═══════════════════════════════════════════════════════════════════════════════
def build_excel(client, product, offer_no, date_obj, project_name,
                rack_data, rate_per_kg, acc_data=None,
                out_path="quotation.xlsx", logo_path=None):
    wb = Workbook()
    ws_q = wb.active
    ws_q.title = "Commercial Offer"
    ws_b = wb.create_sheet("Bill of Materials")

    total_basic, gst, grand = build_quotation_sheet(
        ws_q, client, product, offer_no, date_obj,
        project_name, rack_data, rate_per_kg, acc_data, logo_path
    )
    build_bom_sheet(ws_b, client, offer_no, date_obj, rack_data, acc_data)

    wb.save(out_path)
    return total_basic, gst, grand


# ═══════════════════════════════════════════════════════════════════════════════
#  STREAMLIT UI  —  DSS-forward redesign
# ═══════════════════════════════════════════════════════════════════════════════
st.set_page_config(page_title="DSS Quotation Generator", layout="wide", page_icon="🐬")

st.markdown("""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap');
  html, body, [class*="css"] { font-family: 'Inter', Arial, sans-serif; }

  /* ── Background ── */
  .stApp { background: #eef2f7; }
  section[data-testid="stSidebar"] { display: none; }

  /* ══════════════════════════════════════════════
     HERO  —  DSS is the star
  ══════════════════════════════════════════════ */
  .hero-wrap {
      background: linear-gradient(150deg, #071420 0%, #0D2137 45%, #0a1c30 100%);
      border-radius: 20px;
      margin-bottom: 32px;
      box-shadow: 0 16px 56px rgba(13,33,55,0.50);
      overflow: hidden;
      position: relative;
  }
  .hero-top-stripe {
      height: 7px;
      background: linear-gradient(90deg, #E87722, #ff9a45, #E87722);
  }
  .hero-body {
      padding: 36px 48px 30px;
      display: flex;
      align-items: center;
      gap: 36px;
  }
  .hero-left {
      flex: 0 0 auto;
      border-right: 2px solid rgba(232,119,34,0.35);
      padding-right: 36px;
  }
  .hero-dss-text {
      font-size: 5.8rem;
      font-weight: 900;
      color: #E87722;
      letter-spacing: 10px;
      line-height: 1;
      margin: 0;
      text-shadow: 0 0 60px rgba(232,119,34,0.30);
  }
  .hero-right {
      flex: 1;
      text-align: left;
  }
  .hero-brand-full {
      font-size: 1.35rem;
      font-weight: 700;
      color: #FFFFFF;
      letter-spacing: 4px;
      text-transform: uppercase;
      margin: 0 0 8px;
  }
  .hero-company {
      font-size: 0.88rem;
      font-weight: 500;
      color: #8AAECF;
      letter-spacing: 2px;
      text-transform: uppercase;
      margin: 0 0 14px;
  }
  .hero-tag-row {
      display: flex;
      gap: 10px;
      flex-wrap: wrap;
  }
  .hero-tag {
      display: inline-block;
      background: rgba(232,119,34,0.15);
      border: 1px solid rgba(232,119,34,0.40);
      color: #E87722;
      font-size: 0.70rem;
      font-weight: 700;
      letter-spacing: 2px;
      padding: 5px 14px;
      border-radius: 20px;
      text-transform: uppercase;
  }
  .hero-tag-blue {
      background: rgba(46,109,164,0.20);
      border: 1px solid rgba(46,109,164,0.40);
      color: #7eb8e8;
  }
  .hero-bottom-stripe {
      height: 5px;
      background: linear-gradient(90deg, #E87722, #ff9a45, #E87722);
  }

  /* ══════════════════════════════════════════════
     SECTION HEADERS
  ══════════════════════════════════════════════ */
  h2 {
      color: #0D2137 !important;
      font-weight: 800 !important;
      letter-spacing: 0.3px !important;
      font-size: 1.05rem !important;
  }
  .section-header {
      display: flex;
      align-items: center;
      gap: 10px;
      margin-bottom: 16px;
      padding-bottom: 10px;
      border-bottom: 2px solid #E87722;
  }
  .section-header-text {
      font-size: 1.05rem;
      font-weight: 800;
      color: #0D2137;
      letter-spacing: 0.3px;
  }
  .section-pill {
      background: #0D2137;
      color: #E87722;
      font-size: 0.65rem;
      font-weight: 700;
      letter-spacing: 1.5px;
      padding: 3px 10px;
      border-radius: 20px;
      text-transform: uppercase;
  }

  /* ══════════════════════════════════════════════
     CARDS
  ══════════════════════════════════════════════ */
  .glass-card {
      background: #ffffff;
      border-radius: 16px;
      padding: 24px 26px;
      margin-bottom: 20px;
      box-shadow: 0 2px 16px rgba(13,33,55,0.07);
      border-top: 4px solid #E87722;
  }

  /* ══════════════════════════════════════════════
     INPUTS
  ══════════════════════════════════════════════ */
  input, select, textarea {
      border-radius: 10px !important;
      border-color: #C8DFF0 !important;
  }
  input:focus, select:focus {
      border-color: #E87722 !important;
      box-shadow: 0 0 0 3px rgba(232,119,34,0.16) !important;
  }
  label {
      color: #1A3A5C !important;
      font-weight: 600 !important;
      font-size: 0.82rem !important;
      letter-spacing: 0.2px !important;
  }

  /* ══════════════════════════════════════════════
     EXPANDERS
  ══════════════════════════════════════════════ */
  .streamlit-expanderHeader {
      background: #EAF2FB !important;
      border-left: 4px solid #E87722 !important;
      border-radius: 12px !important;
      font-weight: 700 !important;
      color: #0D2137 !important;
      font-size: 0.90rem !important;
  }
  .streamlit-expanderContent {
      background: #f6f9fd !important;
      border-left: 4px solid rgba(232,119,34,0.20) !important;
      border-radius: 0 0 12px 12px !important;
  }

  /* ══════════════════════════════════════════════
     METRIC CARDS
  ══════════════════════════════════════════════ */
  [data-testid="metric-container"] {
      background: linear-gradient(135deg, #ffffff 0%, #EAF2FB 100%);
      border: 1px solid #C8DFF0;
      border-top: 4px solid #E87722;
      border-radius: 14px;
      padding: 18px 16px !important;
      box-shadow: 0 4px 16px rgba(13,33,55,0.08);
      transition: transform 0.15s ease, box-shadow 0.15s ease;
  }
  [data-testid="metric-container"]:hover {
      transform: translateY(-3px);
      box-shadow: 0 8px 24px rgba(13,33,55,0.14);
  }
  [data-testid="metric-container"] label {
      color: #1A3A5C !important;
      font-weight: 700 !important;
      font-size: 0.75rem !important;
      text-transform: uppercase !important;
      letter-spacing: 0.5px !important;
  }
  [data-testid="metric-container"] [data-testid="metric-value"] {
      color: #0D2137 !important;
      font-size: 1.15rem !important;
      font-weight: 900 !important;
  }

  /* ══════════════════════════════════════════════
     LIVE PREVIEW BAR
  ══════════════════════════════════════════════ */
  .preview-bar {
      background: linear-gradient(135deg, #0D2137 0%, #1A3A5C 100%);
      border-radius: 14px;
      padding: 14px 24px;
      margin: 4px 0 20px;
      border-left: 5px solid #E87722;
      display: flex;
      align-items: center;
      gap: 12px;
  }
  .preview-bar-label {
      color: #E87722;
      font-size: 0.75rem;
      font-weight: 800;
      letter-spacing: 2px;
      text-transform: uppercase;
      margin: 0;
  }
  .preview-bar-sub {
      color: #6A8FAA;
      font-size: 0.72rem;
      margin: 2px 0 0;
  }

  /* ══════════════════════════════════════════════
     COLUMN LABELS
  ══════════════════════════════════════════════ */
  .col-label {
      background: linear-gradient(135deg, #0D2137, #1A3A5C);
      color: #E87722;
      font-size: 0.68rem;
      font-weight: 800;
      letter-spacing: 1.5px;
      text-transform: uppercase;
      padding: 5px 13px;
      border-radius: 8px;
      display: inline-block;
      margin-bottom: 12px;
  }

  /* ══════════════════════════════════════════════
     ACCESSORIES BANNER
  ══════════════════════════════════════════════ */
  .acc-banner {
      background: linear-gradient(90deg, #EAF2FB 0%, #f8fbff 100%);
      border-left: 5px solid #E87722;
      padding: 12px 20px;
      border-radius: 10px;
      margin-bottom: 16px;
      color: #1A3A5C;
      font-weight: 500;
      font-size: 0.86rem;
  }

  /* ══════════════════════════════════════════════
     GENERATE BUTTON
  ══════════════════════════════════════════════ */
  .stButton > button {
      background: linear-gradient(135deg, #0D2137 0%, #1A3A5C 100%) !important;
      color: #E87722 !important;
      font-weight: 800 !important;
      border: 2px solid #E87722 !important;
      border-radius: 14px !important;
      padding: 18px 32px !important;
      font-size: 1.05rem !important;
      letter-spacing: 2px !important;
      transition: all 0.22s ease !important;
      box-shadow: 0 6px 24px rgba(13,33,55,0.30) !important;
  }
  .stButton > button:hover {
      background: linear-gradient(135deg, #E87722 0%, #B85C0A 100%) !important;
      color: #FFFFFF !important;
      transform: translateY(-3px) !important;
      box-shadow: 0 10px 32px rgba(232,119,34,0.50) !important;
      border-color: transparent !important;
  }

  /* ══════════════════════════════════════════════
     DOWNLOAD BUTTON
  ══════════════════════════════════════════════ */
  [data-testid="stDownloadButton"] button {
      background: linear-gradient(135deg, #E87722 0%, #B85C0A 100%) !important;
      color: white !important;
      border: none !important;
      border-radius: 14px !important;
      font-weight: 800 !important;
      font-size: 1.02rem !important;
      padding: 16px 28px !important;
      box-shadow: 0 6px 22px rgba(232,119,34,0.45) !important;
      transition: all 0.22s ease !important;
  }
  [data-testid="stDownloadButton"] button:hover {
      background: linear-gradient(135deg, #ff9a45 0%, #E87722 100%) !important;
      transform: translateY(-3px) !important;
      box-shadow: 0 10px 30px rgba(232,119,34,0.60) !important;
  }

  /* ══════════════════════════════════════════════
     ALERTS & DIVIDER
  ══════════════════════════════════════════════ */
  hr { border-color: #E87722 !important; border-width: 1.5px !important; opacity: 0.18; }
  .stSuccess {
      border-left: 5px solid #E87722 !important;
      background: #FEF0E4 !important;
      border-radius: 12px !important;
  }
  .stInfo    { border-left: 5px solid #1A3A5C !important; border-radius: 12px !important; }
  .stError   { border-radius: 12px !important; }
</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
#  HERO  —  DSS front and centre
# ═══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<div class="hero-wrap">
  <div class="hero-top-stripe"></div>
  <div class="hero-body">
    <div class="hero-left">
      <div class="hero-dss-text">DSS</div>
    </div>
    <div class="hero-right">
      <div class="hero-brand-full">Dolphin Storage Solutions</div>
      <div class="hero-company">🏭 &nbsp; BRIJ INDUSTRIES &nbsp;·&nbsp; Mundka Industrial Area, New Delhi</div>
      <div class="hero-tag-row">
        <span class="hero-tag">Quotation Generator</span>
        <span class="hero-tag hero-tag-blue">Modular Racking Systems</span>
        <span class="hero-tag hero-tag-blue">Mezzanine Floors</span>
      </div>
    </div>
  </div>
  <div class="hero-bottom-stripe"></div>
</div>
""", unsafe_allow_html=True)

# ── Logo upload ───────────────────────────────────────────────────────────────
with st.expander("🖼️  Upload Company Logo  (optional)", expanded=False):
    logo_file = st.file_uploader("Upload logo PNG / JPG", type=["png", "jpg", "jpeg"])

logo_path = None
if logo_file:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tf:
        tf.write(logo_file.read())
        logo_path = tf.name
    st.success(f"✅  Logo uploaded: {logo_file.name}")

# ── Customer & Offer Details ──────────────────────────────────────────────────
st.markdown("""
<div class="section-header">
  <span class="section-header-text">📋 Customer & Offer Details</span>
  <span class="section-pill">Step 1</span>
</div>
""", unsafe_allow_html=True)

c1, c2, c3 = st.columns(3)
with c1:
    st.markdown('<span class="col-label">Client Info</span>', unsafe_allow_html=True)
    client       = st.text_input("Customer Name (M/S)", "STYLE BAZAAR")
    product      = st.text_input("Product",             "MODULAR MEZZANINE FLOOR")
with c2:
    st.markdown('<span class="col-label">Offer Info</span>', unsafe_allow_html=True)
    offer_no     = st.text_input("Offer No",    "DSS-IV/25-26/0712")
    project_name = st.text_input("Project Name","MODULE MEZZANINE FLOOR")
with c3:
    st.markdown('<span class="col-label">Pricing</span>', unsafe_allow_html=True)
    date        = st.date_input("Date", datetime.date.today())
    rate_per_kg = st.number_input("Rate per KG (Rs.)", value=85.00, min_value=0.0, format="%.2f")

st.divider()

# ── Rack Configurations ───────────────────────────────────────────────────────
st.markdown("""
<div class="section-header">
  <span class="section-header-text">🏗️ Rack Configurations</span>
  <span class="section-pill">Step 2</span>
</div>
""", unsafe_allow_html=True)

rack_types = st.number_input("Number of Rack Types", min_value=1, max_value=10, value=1)

rack_data = []
for i in range(int(rack_types)):
    default_name = f"Module {chr(65 + i)}"
    with st.expander(f"⚙️  Rack Type {i+1}  —  {default_name}", expanded=(i == 0)):
        module_name = st.text_input(
            "Module Name / Label", value=default_name, key=f"mn{i}",
            help="E.g. 'Module A', 'Ground Floor', 'Block 1'"
        )

        c1, c2, c3, c4, c5 = st.columns(5)
        with c1:
            st.markdown('<span class="col-label">Quantities</span>', unsafe_allow_html=True)
            main_qty  = st.number_input("Main Rack Qty",   key=f"mq{i}", value=10, min_value=0)
            addon_qty = st.number_input("Add-on Rack Qty", key=f"aq{i}", value=5,  min_value=0)
            levels    = st.number_input("No. of Levels",   key=f"lv{i}", value=3,  min_value=1)

        with c2:
            st.markdown('<span class="col-label">Upright</span>', unsafe_allow_html=True)
            upright_section = st.selectbox(
                "Section Type",
                ["Box Section", "Omega Section 70", "Omega Section 90"],
                key=f"us{i}", index=1
            )
            if upright_section == "Box Section":
                uw = st.number_input("Width (mm)",  key=f"uw{i}", value=80)
                ud = st.number_input("Depth (mm)",  key=f"ud{i}", value=60)
            elif upright_section == "Omega Section 90":
                uw = 70; ud = 90
                st.info("70 × 90 mm Omega")
            else:
                uw = 70; ud = 110
                st.info("70 × 110 mm Omega")
            ul = st.number_input("Height (mm)",    key=f"ul{i}", value=3000)
            ut = st.number_input("Thickness (mm)", key=f"ut{i}", value=1.6, format="%.1f")

        with c3:
            st.markdown('<span class="col-label">Beam</span>', unsafe_allow_html=True)
            bt  = st.selectbox("Beam Type", ["Pipe Beam", "Roll Formed Beam"], key=f"bt{i}")
            bh  = st.number_input("Height (mm)",    key=f"bh{i}", value=100)
            bw  = st.number_input("Width (mm)",     key=f"bw{i}", value=50)
            bl  = st.number_input("Length (mm)",    key=f"bl{i}", value=2000)
            bth = st.number_input("Thickness (mm)", key=f"bth{i}", value=1.6, format="%.1f")

        with c4:
            st.markdown('<span class="col-label">Rack / Deep Bar</span>', unsafe_allow_html=True)
            depth = st.number_input("Rack Depth (mm)",     key=f"dp{i}", value=800)
            dth   = st.number_input("Deep Bar Thick (mm)", key=f"dt{i}", value=1.6, format="%.1f")

        with c5:
            st.markdown('<span class="col-label">Cross Brace</span>', unsafe_allow_html=True)
            method = st.number_input("Method (mm)", min_value=200, max_value=500, step=1, key=f"mt{i}")
            gap    = st.number_input("Gap (mm)",    min_value=600, max_value=900, step=1, key=f"gp{i}")
            cth    = st.number_input("Cross Thick (mm)", key=f"ct{i}", value=1.6, format="%.1f")

        rack_data.append({
            "module": module_name,
            "main_qty": main_qty, "addon_qty": addon_qty, "levels": levels,
            "uw": uw, "ud": ud, "ul": ul, "ut": ut,
            "bt": bt, "bh": bh, "bw": bw, "bl": bl, "bth": bth,
            "depth": depth, "dth": dth,
            "method": method, "gap": gap, "cth": cth,
        })

st.divider()

# ── Accessories ───────────────────────────────────────────────────────────────
st.markdown("""
<div class="section-header">
  <span class="section-header-text">🔩 Accessories</span>
  <span class="section-pill">Step 3</span>
</div>
""", unsafe_allow_html=True)

st.markdown(
    '<div class="acc-banner">Enter quantities and dimensions for each accessory. '
    'Items with <strong>Qty = 0</strong> are automatically excluded from the quotation.</div>',
    unsafe_allow_html=True
)

with st.expander("⚙️  Configure Accessories", expanded=False):
    a1, a2 = st.columns(2)
    with a1:
        st.markdown('<span class="col-label">Standard Items</span>', unsafe_allow_html=True)
        cg_qty = st.number_input("Column Guard Qty",  min_value=0, value=0, key="cg")
        rc_qty = st.number_input("Row Connector Qty", min_value=0, value=0, key="rc")
    with a2:
        st.markdown('<span class="col-label">Row Guards</span>', unsafe_allow_html=True)
        rg_types_n = st.number_input("Row Guard Types", min_value=0, max_value=5, value=0, key="rgt")

    row_guards = []
    for j in range(int(rg_types_n)):
        with st.expander(f"Row Guard — Type {j+1}"):
            rg_c1, rg_c2 = st.columns(2)
            rg_l   = rg_c1.number_input("Length (mm)", key=f"rgl{j}", value=2000.0, format="%.1f")
            rg_qty = rg_c2.number_input("Qty",         key=f"rgq{j}", value=1, min_value=0)
            st.caption(f"Row Guard Type {j+1}:  L = {rg_l:.0f} mm  |  Qty = {rg_qty}")
            row_guards.append({"h": 400, "l": rg_l, "qty": rg_qty})

    st.markdown('<span class="col-label">Tie Beams</span>', unsafe_allow_html=True)
    tb_types_n = st.number_input("Tie Beam Types", min_value=0, max_value=5, value=0, key="tbt")
    tie_beams = []
    for j in range(int(tb_types_n)):
        with st.expander(f"Tie Beam — Type {j+1}"):
            tc1, tc2, tc3 = st.columns(3)
            tb_qty = tc1.number_input("Qty",          key=f"tbq{j}", value=1, min_value=0)
            tb_l   = tc2.number_input("Length (mm)",  key=f"tbl{j}", value=2000.0, format="%.1f")
            tb_t   = tc3.number_input("Thickness (mm)", key=f"tbt2{j}", value=1.6, format="%.1f")
            st.caption(f"Tie Beam Type {j+1}:  L = {tb_l:.0f} mm  |  t = {tb_t} mm  |  Qty = {tb_qty}")
            tie_beams.append({"qty": tb_qty, "w": 80.0, "d": 60.0, "l": tb_l, "t": tb_t})

    st.markdown('<span class="col-label">Back Pallet Stoppers</span>', unsafe_allow_html=True)
    bps_types_n = st.number_input("BPS Types", min_value=0, max_value=5, value=0, key="bpst")
    bps_list = []
    for j in range(int(bps_types_n)):
        with st.expander(f"Back Pallet Stopper — Type {j+1}"):
            bc1, bc2 = st.columns(2)
            bps_qty = bc1.number_input("Qty",         key=f"bpsq{j}", value=1, min_value=0)
            bps_l   = bc2.number_input("Length (mm)", key=f"bpsl{j}", value=2000.0, format="%.1f")
            st.caption(f"BPS Type {j+1}:  L = {bps_l:.0f} mm  |  Qty = {bps_qty}")
            bps_list.append({"qty": bps_qty, "l": bps_l})

acc_data = {
    "cg_qty": cg_qty, "rc_qty": rc_qty,
    "row_guards": row_guards, "tie_beams": tie_beams, "bps_list": bps_list,
}

st.divider()

# ── Live Preview ──────────────────────────────────────────────────────────────
if rack_data:
    comp0  = calc_components(rack_data[0])
    all_wt = sum(
        calc_components(r)["total_main"]  * r["main_qty"] +
        calc_components(r)["total_addon"] * r["addon_qty"]
        for r in rack_data
    )
    prev_tot   = all_wt * rate_per_kg
    prev_gst   = prev_tot * 0.18
    prev_grand = prev_tot + prev_gst

    st.markdown("""
    <div class="preview-bar">
      <div>
        <p class="preview-bar-label">📊 &nbsp; Live Estimate Preview</p>
        <p class="preview-bar-sub">Updates as you change inputs</p>
      </div>
    </div>
    """, unsafe_allow_html=True)

    p1, p2, p3, p4, p5, p6 = st.columns(6)
    p1.metric("Main Rack Wt (Mod 1)",   f"{comp0['total_main']:.3f} kg")
    p2.metric("Add-on Wt (Mod 1)",      f"{comp0['total_addon']:.3f} kg")
    p3.metric("Total Tonnage",          f"{all_wt/1000:.3f} MT")
    p4.metric("Basic Amount",           f"Rs. {prev_tot:,.0f}")
    p5.metric("GST  (18%)",             f"Rs. {prev_gst:,.0f}")
    p6.metric("Grand Total",            f"Rs. {prev_grand:,.0f}")

st.divider()

# ── Generate Button ───────────────────────────────────────────────────────────
st.markdown("""
<div class="section-header">
  <span class="section-header-text">🐬 Generate Documents</span>
  <span class="section-pill">Step 4</span>
</div>
""", unsafe_allow_html=True)

if st.button("🐬  GENERATE  QUOTATION  +  BOM", type="primary", use_container_width=True):
    safe  = client.replace(" ", "_").replace("/", "-")
    fname = f"DSS_{safe}_Offer_{offer_no.replace('/', '-')}.xlsx"
    out   = os.path.join(tempfile.gettempdir(), fname)

    try:
        basic, gst_amt, grand = build_excel(
            client, product, offer_no, date, project_name,
            rack_data, rate_per_kg, acc_data=acc_data,
            out_path=out, logo_path=logo_path
        )
        st.success("✅  DSS Workbook ready — Commercial Offer + Bill of Materials sheets generated.")
        with open(out, "rb") as f:
            st.download_button(
                "⬇️  DOWNLOAD  DSS EXCEL  (Quotation + BOM)",
                data=f, file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        r1, r2, r3 = st.columns(3)
        r1.metric("Basic Amount",  f"Rs. {basic:,.2f}")
        r2.metric("GST (18%)",     f"Rs. {gst_amt:,.2f}")
        r3.metric("Grand Total",   f"Rs. {grand:,.2f}", delta="Incl. GST")
    except Exception as e:
        st.error(f"❌  Error generating file: {e}")
        raise
